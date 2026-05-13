import re
import os
import openpyxl
from pyxlsb import open_workbook


def build_product_mapping(source_path):
    """从 For Mapping sheet 读取 商品编号 → (品牌, 单盒/大包装)"""
    mapping = {}
    wb = open_workbook(source_path)

    sheet_name = None
    for s in wb.sheets:
        if 'For Mapping' in s or 'for mapping' in s.lower():
            sheet_name = s
            break

    if not sheet_name:
        print("错误: 未找到 For Mapping sheet")
        wb.close()
        return None

    print(f"使用sheet: {sheet_name}")

    with wb.get_sheet(sheet_name) as ws:
        for i, row in enumerate(ws):
            if i < 2:
                continue

            # 左侧京东产品Mapping
            sku_left = row[0].v
            brand_left = row[2].v if row[2].v else ""
            pkg_left = row[3].v if row[3].v else ""
            if sku_left is not None:
                sku = str(int(sku_left)) if isinstance(sku_left, float) else str(sku_left)
                mapping[sku] = (str(brand_left), str(pkg_left))

            # 右侧千牛产品Mapping
            sku_right = row[10].v if len(row) > 10 else None
            brand_right = row[12].v if len(row) > 12 and row[12].v else ""
            pkg_right = row[13].v if len(row) > 13 and row[13].v else ""
            if sku_right is not None:
                sku = str(int(sku_right)) if isinstance(sku_right, float) else str(sku_right)
                if sku not in mapping:
                    mapping[sku] = (str(brand_right), str(pkg_right))

    wb.close()
    print(f"商品映射构建完成，共 {len(mapping)} 条")
    return mapping


def extract_sku_from_link(link):
    if not link:
        return ""
    m = re.search(r'item\.jd\.com/(\d+)', link)
    return m.group(1) if m else ""


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()


def add_product_info_to_sheet(ws, mapping):
    """为单个sheet添加 商品编号, 品牌, 单盒/大包装 列"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {cell: idx for idx, cell in enumerate(header_row)}

    if '进线链接' not in headers:
        return None
    if '商品编号' in headers:
        return None

    link_col = headers['进线链接']  # 0-based

    # 在进线链接列右侧插入3列（openpyxl 1-based）
    insert_at = link_col + 2  # 1-based, 进线链接列的下一列
    ws.insert_cols(insert_at, 3)

    # 从进线链接列复制表头样式到新列
    ref_header = ws.cell(row=1, column=link_col + 1)  # 1-based
    for col_offset, title in enumerate(['商品编号', '品牌', '单盒/大包装']):
        new_cell = ws.cell(row=1, column=insert_at + col_offset, value=title)
        copy_cell_style(ref_header, new_cell)

    matched = 0
    unmatched = 0
    total = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if all(cell.value is None for cell in row):
            continue

        link_cell = row[link_col]
        link = str(link_cell.value) if link_cell.value else ""
        sku = extract_sku_from_link(link)
        ref_data_cell = row[link_col]

        # 写入商品编号
        sku_cell = ws.cell(row=row_idx, column=insert_at, value=sku)
        copy_cell_style(ref_data_cell, sku_cell)

        if sku and sku in mapping:
            brand, pkg = mapping[sku]
            brand_cell = ws.cell(row=row_idx, column=insert_at + 1, value=brand)
            pkg_cell = ws.cell(row=row_idx, column=insert_at + 2, value=pkg)
            copy_cell_style(ref_data_cell, brand_cell)
            copy_cell_style(ref_data_cell, pkg_cell)
            matched += 1
        else:
            unmatched += 1
        total += 1

    return (matched, unmatched, total)


def main():
    print("=" * 60)
    print("为客服数据添加商品信息")
    print("=" * 60)

    SOURCE_FILE = input("请输入For Mapping来源文件路径: ").strip().strip('"')
    INPUT_FILE = input("请输入要添加商品信息的目标文件路径: ").strip().strip('"')

    if not os.path.exists(SOURCE_FILE):
        print(f"错误: 源文件不存在 {SOURCE_FILE}")
        return
    if not os.path.exists(INPUT_FILE):
        print(f"错误: 文件不存在 {INPUT_FILE}")
        return

    mapping = build_product_mapping(SOURCE_FILE)
    if mapping is None:
        return

    wb = openpyxl.load_workbook(INPUT_FILE)

    total_matched = total_unmatched = total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result = add_product_info_to_sheet(ws, mapping)
        if result is None:
            print(f"  {sheet_name}: 跳过（无进线链接或已有商品编号）")
        else:
            m, u, t = result
            total_matched += m
            total_unmatched += u
            total_rows += t
            print(f"  {sheet_name}: {t}行, 匹配{m}, 未匹配{u}")

    temp_file = INPUT_FILE.replace('.xlsx', '_temp.xlsx')
    wb.save(temp_file)
    wb.close()

    try:
        os.replace(temp_file, INPUT_FILE)
    except:
        if os.path.exists(INPUT_FILE):
            os.remove(INPUT_FILE)
        os.rename(temp_file, INPUT_FILE)

    print(f"\n处理完成: 共 {total_rows} 行")
    print(f"  匹配到商品信息: {total_matched} 行")
    print(f"  未匹配:         {total_unmatched} 行")
    print(f"文件已保存到: {INPUT_FILE}")


if __name__ == "__main__":
    main()
