import re
import os
import openpyxl


def convert_packaging(name, packaging):
    """根据商品名称和原包装类型，转换单盒/大包装为具体规格"""
    if packaging == '单盒':
        # 找 XX片，排除 XX片/板、XX片*板，但允许 XX片/盒、XX片*盒
        matches = re.findall(r'(\d+)片(?![\*/]板)', name)
        if matches:
            return f"1盒{matches[-1]}片"

    elif packaging == '大包装':
        # 优先匹配 XX片/板*Y板
        m = re.search(r'(\d+)片[/\*]板[\*/]?(\d+)板', name)
        if m:
            return f"{m.group(2)}盒{m.group(1)}片"
        # XX片*Y板
        m = re.search(r'(\d+)片[\*/](\d+)板', name)
        if m:
            return f"{m.group(2)}盒{m.group(1)}片"
        # XX片 → 1盒XX片
        matches = re.findall(r'(\d+)片', name)
        if matches:
            return f"1盒{matches[-1]}片"

    return packaging


def main():
    INPUT_FILE = r"D:\AIwork\test\month\mapping修改.xlsx"

    print("=" * 60)
    print("修改单盒/大包装列")
    print("=" * 60)

    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {cell: idx for idx, cell in enumerate(header)}
    print(f"表头: {header}")

    pkg_col = headers.get('单盒/大包装')
    name_col = headers.get('商品名称')
    if pkg_col is None or name_col is None:
        print("错误: 找不到列")
        wb.close()
        return

    modified = 0
    unchanged = 0

    for row in ws.iter_rows(min_row=2):
        name_cell = row[name_col]
        pkg_cell = row[pkg_col]

        if name_cell.value is None:
            continue

        name = str(name_cell.value)
        old_pkg = str(pkg_cell.value) if pkg_cell.value else ""
        new_pkg = convert_packaging(name, old_pkg)

        if new_pkg != old_pkg:
            pkg_cell.value = new_pkg
            modified += 1
            print(f"  {old_pkg} → {new_pkg}  ({name[:60]}...)")
        else:
            unchanged += 1

    temp_file = INPUT_FILE.replace('.xlsx', '_temp.xlsx')
    wb.save(temp_file)
    wb.close()

    try:
        os.replace(temp_file, INPUT_FILE)
    except:
        if os.path.exists(INPUT_FILE):
            os.remove(INPUT_FILE)
        os.rename(temp_file, INPUT_FILE)

    print(f"\n修改完成: {modified} 行已修改, {unchanged} 行未变")
    print(f"文件已保存到: {INPUT_FILE}")


if __name__ == "__main__":
    main()
