import os
import openpyxl
from collections import defaultdict
from pyxlsb import open_workbook


def normalize_id(val):
    """处理数字类型ID，避免 13911832121.0"""
    if isinstance(val, float):
        return str(int(val))
    return str(val).strip() if val else ""


def time_to_seconds(t):
    """HH:MM:SS → 秒数"""
    parts = t.split(":")
    return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])


def build_cid_lookup(source_path, sheet_keyword):
    """从源文件读取CID映射。
    返回:
        second_lookup: (日期, HH:MM:SS, 客服, 顾客) → cid
        minute_pool: (日期, HH:MM, 客服, 顾客) → [cid, ...]
        fuzzy_pool: (日期, 客服, 顾客) → [(秒数, cid), ...]  用于±5秒模糊匹配
    """
    second_lookup = {}
    minute_pool = defaultdict(list)
    fuzzy_pool = defaultdict(list)

    wb = open_workbook(source_path)
    target_sheet = None
    for s in wb.sheets:
        if sheet_keyword in s:
            target_sheet = s
            break

    if not target_sheet:
        print(f"错误: 未找到包含 '{sheet_keyword}' 的sheet")
        wb.close()
        return None, None, None

    print(f"使用sheet: {target_sheet}")

    with wb.get_sheet(target_sheet) as ws:
        for i, row in enumerate(ws):
            if i == 0:
                continue

            consult_time = row[2].v if row[2].v else ""
            agent = normalize_id(row[5].v) if row[5].v else ""
            customer = normalize_id(row[6].v) if row[6].v else ""
            cid = normalize_id(row[24].v) if row[24].v else ""

            if not consult_time or not cid:
                continue

            consult_time = str(consult_time).strip().lstrip("'")
            parts = consult_time.split(" ")
            if len(parts) < 2:
                continue
            start_date = parts[0].replace("-", "/")
            full_time = parts[1]

            if not agent or not customer:
                continue

            # 秒级
            sec_key = (start_date, full_time, agent, customer)
            second_lookup[sec_key] = cid

            # 分钟级CID池
            min_key = (start_date, full_time[:5], agent, customer)
            minute_pool[min_key].append(cid)

            # 模糊匹配池: (日期, 客服, 顾客) → (秒数, cid)
            fuzzy_key = (start_date, agent, customer)
            fuzzy_pool[fuzzy_key].append((time_to_seconds(full_time), cid))

    wb.close()
    print(f"秒级映射: {len(second_lookup)} 条, 分钟级CID池: {len(minute_pool)} 组, 模糊池: {len(fuzzy_pool)} 组")
    return second_lookup, minute_pool, fuzzy_pool


def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = source_cell.font.copy()
        target_cell.border = source_cell.border.copy()
        target_cell.fill = source_cell.fill.copy()
        target_cell.number_format = source_cell.number_format
        target_cell.protection = source_cell.protection.copy()
        target_cell.alignment = source_cell.alignment.copy()


def add_cid_to_sheet(ws, second_lookup, minute_pool, fuzzy_pool, used_cids):
    """为单个sheet添加CID列，返回统计信息"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {cell: idx for idx, cell in enumerate(header_row)}

    # 检查是否有需要的列
    required = ['开始日期', '开始时间', '客服姓名', '用户ID']
    for col in required:
        if col not in headers:
            return None

    # 检查是否已有CID列
    if 'CID' in headers:
        return None

    # 在最后一列后添加CID表头，复制前一列样式
    cid_col = len(header_row) + 1
    ref_header = ws.cell(row=1, column=cid_col - 1)
    cid_header = ws.cell(row=1, column=cid_col, value='CID')
    copy_cell_style(ref_header, cid_header)

    matched_sec = 0
    matched_min = 0
    matched_fuzzy = 0
    unmatched = 0
    total = 0

    # 暂存行数据用于模糊匹配
    rows_data = []

    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue

        start_date = str(row[headers['开始日期']].value or "")
        full_time = str(row[headers['开始时间']].value or "")
        minute_time = full_time[:5] if full_time else ""
        agent = str(row[headers['客服姓名']].value or "")
        user_id_raw = row[headers['用户ID']].value
        user_id = normalize_id(user_id_raw)

        cid = ""

        sec_key = (start_date, full_time, agent, user_id)
        if sec_key in second_lookup:
            cid = second_lookup[sec_key]
            matched_sec += 1
        else:
            min_key = (start_date, minute_time, agent, user_id)
            if min_key in minute_pool:
                for pool_cid in minute_pool[min_key]:
                    if pool_cid not in used_cids:
                        cid = pool_cid
                        matched_min += 1
                        break

        if cid:
            used_cids.add(cid)
        else:
            unmatched += 1

        ref_cell = row[0]  # 用第一列样式作为参考
        cid_cell = ws.cell(row=row[0].row, column=cid_col, value=cid)
        copy_cell_style(ref_cell, cid_cell)
        rows_data.append((row[0].row, cid, start_date, full_time, agent, user_id))
        total += 1

    # ±5秒模糊匹配
    for row_num, cid, start_date, full_time, agent, user_id in rows_data:
        if cid:
            continue
        if not full_time or not start_date:
            continue

        input_seconds = time_to_seconds(full_time)
        fuzzy_key = (start_date, agent, user_id)

        if fuzzy_key not in fuzzy_pool:
            continue

        for src_seconds, src_cid in fuzzy_pool[fuzzy_key]:
            if abs(src_seconds - input_seconds) <= 5 and src_cid not in used_cids:
                ref_cell = ws.cell(row=row_num, column=1)
                cid_cell = ws.cell(row=row_num, column=cid_col, value=src_cid)
                copy_cell_style(ref_cell, cid_cell)
                used_cids.add(src_cid)
                matched_fuzzy += 1
                unmatched -= 1
                break

    return (matched_sec, matched_min, matched_fuzzy, unmatched, total)


def process_excel(input_file, output_file, second_lookup, minute_pool, fuzzy_pool):
    print(f"\n读取输入文件: {input_file}")
    wb = openpyxl.load_workbook(input_file)

    used_cids = set()
    total_sec = total_min = total_fuzzy = total_unmatched = total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result = add_cid_to_sheet(ws, second_lookup, minute_pool, fuzzy_pool, used_cids)
        if result is None:
            print(f"  {sheet_name}: 跳过（无需处理或已有CID）")
        else:
            ms, mm, mf, um, tr = result
            total_sec += ms
            total_min += mm
            total_fuzzy += mf
            total_unmatched += um
            total_rows += tr
            print(f"  {sheet_name}: {tr}行, 秒级{ms}, 分钟级{mm}, 模糊{mf}, 未匹配{um}")

    temp_file = output_file.replace('.xlsx', '_temp.xlsx')
    wb.save(temp_file)
    wb.close()

    try:
        os.replace(temp_file, output_file)
    except:
        if os.path.exists(output_file):
            os.remove(output_file)
        os.rename(temp_file, output_file)

    print(f"\n处理完成: 共 {total_rows} 行")
    print(f"  秒级精确匹配: {total_sec} 行")
    print(f"  分钟级分配:   {total_min} 行")
    print(f"  ±5秒模糊匹配: {total_fuzzy} 行")
    print(f"  未匹配:       {total_unmatched} 行")
    print(f"  总匹配:       {total_sec + total_min + total_fuzzy} 行")
    print(f"文件已保存到: {output_file}")


def main():
    print("=" * 60)
    print("为客服数据添加CID列 (秒级优先 + 分钟池 + ±5秒模糊)")
    print("=" * 60)

    SOURCE_FILE = input("请输入咨询会话查询来源文件路径: ").strip().strip('"')
    INPUT_FILE = input("请输入要添加CID的目标文件路径: ").strip().strip('"')
    SHEET_KEYWORD = "咨询会话查询"

    if not os.path.exists(SOURCE_FILE):
        print(f"错误: 源文件不存在 {SOURCE_FILE}")
        return

    if not os.path.exists(INPUT_FILE):
        print(f"错误: 文件不存在 {INPUT_FILE}")
        return

    second_lookup, minute_pool, fuzzy_pool = build_cid_lookup(SOURCE_FILE, SHEET_KEYWORD)
    if second_lookup is None:
        return

    process_excel(INPUT_FILE, INPUT_FILE, second_lookup, minute_pool, fuzzy_pool)
    print("=" * 60)
    print("完成")


if __name__ == "__main__":
    main()
