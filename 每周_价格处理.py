import pandas as pd
import os
from openpyxl.styles import PatternFill

# 1. 路径设置
input_path = r"D:\rpa\data.xlsx"
output_dir = r"D:\rpa"  # 设置输出目录

# 固定店铺列表
fixed_shops = {
    '优思明京东': [
        '海王星辰健康药房旗舰店', '海王星辰大药房旗舰店', '益丰大药房旗舰店', '老百姓大药房官方旗舰店',
        '老百姓大药房旗舰店', '老百姓(衡阳)大药房旗舰店', '老百姓（河南）大药房旗舰店', '老百姓广西大药房旗舰店',
        '老百姓医药专营店', '荷塘月色老百姓大药房旗舰店', '老百姓江南大药房旗舰店', '健之佳官方旗舰店',
        '健之佳大药房旗舰店', '健之佳e购大药房旗舰店', '健之佳佳急送大药房旗舰店', '健之佳（云南）大药房旗舰店',
        '健之佳（重庆）大药房旗舰店', '漱玉（枣庄）大药房旗舰店', '漱玉（临沂）大药房旗舰店', '广药晨菲大药房旗舰店',
        '广药大药房旗舰店', '大参林大药房官方旗舰店', '大参林柏康大药房旗舰店', '国大药房（吉林）大药房旗舰店',
        '国大(广西)大药房旗舰店', '国大药房旗舰店', '国大大药房旗舰店', '南京国大药房旗舰店',
        '国大（武汉）大药房旗舰店', '国大（深圳）大药房旗舰店'
    ],
    '优思明阿里': [
        '海王星辰大药房旗舰店', '益丰大药房旗舰店', '广西老百姓大药房旗舰店', '金华老百姓大药房旗舰店',
        '嘉兴老百姓大药房旗舰店', '海宁老百姓大药房旗舰店', '老百姓大药房旗舰店', '健之佳大药房旗舰店',
        '健之佳e购大药房旗舰店', '健之佳佳急送大药房旗舰店', '重庆健之佳大药房旗舰店', '漱玉平民大药房旗舰店',
        '临沂漱玉平民大药房旗舰店', '枣庄漱玉平民大药房旗舰店', '广药', '大参林大药房旗舰店',
        '国大药房福建大药房旗舰店', '国大药房旗舰店', '国大益和大药房旗舰', '国大药房广西大药房旗舰店',
        '国大药房南京大药房旗舰店', '国大药房宁夏大药房旗舰店', '国大药房沈阳大药房旗舰店', '武汉国大大药房旗舰店',
        '国大深圳大药房旗舰店'
    ],
    '优思悦京东': [
        '海王星辰（江苏）大药房旗舰店','海王星辰大药房旗舰店', '海王星辰四川大药房旗舰店', '海王星辰健康药房旗舰店', '益丰大药房旗舰店',
        '益丰（湖南）大药房旗舰店', '老百姓大药房旗舰店', '老百姓(衡阳)大药房旗舰店', '老百姓（陕西）大药房旗舰店',
        '老百姓（湖北）大药房旗舰店', '桐乡老百姓大药房旗舰店', '老百姓广西大药房旗舰店', '老百姓（河南）大药房旗舰店',
        '老百姓医药专营店', '老百姓大药房官方旗舰店', '健之佳官方旗舰店', '健之佳大药房旗舰店',
        '健之佳e购大药房旗舰店', '健之佳（云南）大药房旗舰店', '健之佳（重庆）大药房旗舰店', '漱玉平民大药房旗舰店',
        '漱玉健康大药房旗舰店', '广药大药房旗舰店', '大参林大药房官方旗舰店', '国大（武汉）大药房旗舰店',
        '国大药房（吉林）大药房旗舰店', '国大（深圳）大药房旗舰店'
    ],
    '优思悦阿里': [
        '杭州海王星辰大药房旗舰店', '海王星辰大药房旗舰店', '大连海王星辰大药房旗舰店', '益丰大药房旗舰店',
        '老百姓大药房旗舰店', '衡阳老百姓大药房旗舰店', '嘉兴老百姓大药房旗舰店', '广西老百姓大药房旗舰店',
        '陕西老百姓大药房旗舰店', '健之佳大药房旗舰店', '健之佳佳急送大药房旗舰店', '重庆健之佳大药房旗舰店',
        '健之佳e购大药房旗舰店', '漱玉平民大药房旗舰店', '枣庄漱玉平民大药房旗舰店', '泰安漱玉平民大药房旗舰店',
        '大参林大药房旗舰店', '国大深圳大药房旗舰店', '国大药房沈阳大药房旗舰店', '国大益和大药房旗舰',
        '国大药房宁夏大药房旗舰店', '国大药房广西大药房旗舰店', '国大万民大药房旗舰店', '国大药房旗舰店',
        '武汉国大大药房旗舰店', '国大药房扬州大药房旗舰店'
    ]
}

try:
    # 加载所有数据
    excel_file = pd.ExcelFile(input_path)
    target_sheets = ['优思明', '优思悦']
    target_platforms = ['京东', '阿里']

    # 汇总所有 Sheet 数据
    all_data_list = []
    for sheet_name in excel_file.sheet_names:
        name_clean = sheet_name.strip()
        if name_clean in target_sheets:
            temp_df = pd.read_excel(input_path, sheet_name=sheet_name)
            temp_df.columns = temp_df.columns.str.strip()
            temp_df['original_sheet'] = name_clean
            all_data_list.append(temp_df)

    if not all_data_list:
        print("未找到指定名称的工作表")
        exit()

    full_df = pd.concat(all_data_list, ignore_index=True)

    # --- 处理日期格式 ---
    date_col = '日期'
    full_df[date_col] = pd.to_datetime(full_df[date_col])

    # --- 【新增逻辑】获取时间区间并生成动态文件名 ---
    min_date_str = full_df[date_col].min().strftime('%m%d')
    max_date_str = full_df[date_col].max().strftime('%m%d')
    dynamic_filename = f"{min_date_str}-{max_date_str}价格汇总.xlsx"
    output_path = os.path.join(output_dir, dynamic_filename)

    # 转换为 MMDD 格式，作为列名
    full_df['日期显示'] = full_df[date_col].dt.strftime('%m%d')

    # 创建一个写入器
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name in target_sheets:
            for platform in target_platforms:
                # 确定当前 sheet 名称
                current_sheet_name = f"{sheet_name}{platform}"

                # 筛选当前药品和平台的全部日期数据
                df = full_df[(full_df['original_sheet'] == sheet_name) & (full_df['平台'] == platform)].copy()

                # 获取当前 sheet 的固定店铺列表
                shops = fixed_shops.get(current_sheet_name, [])

                # 获取所有日期作为列
                if not df.empty:
                    dates = sorted(df['日期显示'].unique())
                else:
                    # 如果没有数据，至少创建一个空的日期列
                    dates = ['']

                # 创建结果数据框
                result_df = pd.DataFrame(columns=['平台', '店铺'] + dates)

                # 填充固定店铺的数据
                row_index = 0
                for shop in shops:
                    # 获取该店铺的所有价格数据
                    shop_data = df[df['店铺'] == shop].copy()

                    if shop_data.empty:
                        # 如果该店铺完全没有数据，创建一行全是 "-"
                        row_data = {'平台': '', '店铺': shop}
                        for date in dates:
                            row_data[date] = '-'

                        result_df.loc[row_index] = row_data
                        row_index += 1
                    else:
                        # 获取该店铺出现过的所有价格
                        prices = sorted(shop_data['价格'].unique())

                        # 为每个价格创建一行
                        for price in prices:
                            row_data = {'平台': '', '店铺': shop}
                            for date in dates:
                                # 检查该店铺在该日期是否有该价格
                                date_price_data = shop_data[(shop_data['日期显示'] == date) & (shop_data['价格'] == price)]
                                if not date_price_data.empty:
                                    row_data[date] = price
                                else:
                                    row_data[date] = '-'

                            result_df.loc[row_index] = row_data
                            row_index += 1

                # 处理不在固定列表中的店铺
                if not df.empty:
                    all_actual_shops = df['店铺'].unique()
                    for shop in all_actual_shops:
                        if shop not in shops:
                            # 获取该店铺的所有价格
                            shop_data = df[df['店铺'] == shop]
                            prices = sorted(shop_data['价格'].unique())

                            for price in prices:
                                row_data = {'平台': '', '店铺': shop}
                                for date in dates:
                                    date_price_data = shop_data[(shop_data['日期显示'] == date) & (shop_data['价格'] == price)]
                                    if not date_price_data.empty:
                                        row_data[date] = price
                                    else:
                                        row_data[date] = '-'  # 改为 '-' 而不是 '未获取'

                                result_df.loc[row_index] = row_data
                                row_index += 1

                # 第一行的平台列填写对应平台名称
                if not result_df.empty:
                    result_df.loc[0, '平台'] = platform

                # 写入对应的 Sheet
                result_df.to_excel(writer, sheet_name=current_sheet_name, index=False)

    # 读取生成的Excel文件，添加样式
    from openpyxl import load_workbook
    wb = load_workbook(output_path)

    # 为每个 sheet 添加黄色底色样式
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 确定当前 sheet 对应的固定店铺列表
        if '优思明京东' in sheet_name:
            current_shops = fixed_shops['优思明京东']
        elif '优思明阿里' in sheet_name:
            current_shops = fixed_shops['优思明阿里']
        elif '优思悦京东' in sheet_name:
            current_shops = fixed_shops['优思悦京东']
        elif '优思悦阿里' in sheet_name:
            current_shops = fixed_shops['优思悦阿里']
        else:
            current_shops = []

        # 遍历除表头外的所有行
        for row in range(2, ws.max_row + 1):
            cell_value = str(ws.cell(row, 2).value) if ws.cell(row, 2).value is not None else ''
            if cell_value not in current_shops:
                # 给不在固定列表中的店铺填充黄色底色
                ws.cell(row, 2).fill = yellow_fill

    # 保存修改
    wb.save(output_path)

    print(f"成功生成汇总文件: {output_path}")

except Exception as e:
    print(f"运行出错: {e}")
    import traceback
    traceback.print_exc()
