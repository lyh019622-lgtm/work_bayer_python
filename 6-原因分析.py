#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
方案C：用户消息优先（单话题）
- 逐条匹配用户消息，取第一个命中
- 用户消息全没命中 → 匹配客服消息，取第一个命中
- 都没命中 → 全对话兜底
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import os
import openpyxl

# ======================== 关键词规则 ========================
KEYWORD_RULES = [
    ("产品质量咨询", "疑似售后-其他",
     ["售后", "转小晟", "转客服", "差评", "投诉", "服务不好","转小象","转小文","转西西","转小佳","转小小","转小洋","转小云","转小宇",
      "快递不好", "评价问题", "不满意", "查订单", "改地址", "取件码", "奖品发货",
      "催奖品发货", "中奖发货"],
     "其他售后问题"),
    ("购买咨询", "价格问题-咨询医保/商保",
     ["商保", "医保支付", "医保统筹", "商保报销", "医保卡", "报不了医保","用医保","医保购买","可以医保","报销","医保","药省保","京东宝","京东好医保"],
     "咨询医保/商保支付"),
    ("医学咨询", "服用方法-补服漏服",
     ["漏服", "补服", "忘记吃", "下一板","开始喝","才吃","补救","忘吃","晚吃了","晚了50分钟","按顺序","分周","今天吃","晚个几小时",
     "忘记吃","吃药","差半小时","差一小时","连续吃","多吃了一片","多吃一片","少吃了一片","少吃一片","30吃的","45吃的","点吃的"],
     "咨询漏服补服方法"),
    ("医学咨询", "服用方法-推迟月经",
     ["推迟月经", "推迟经期", "月经","姨妈","7天"],
     "咨询推迟月经吃法"),
    ("医学咨询", "服用方法-规律经期",
     ["调经", "调理月经", "规律经期", "调节激素", "周期调整"],
     "咨询调理月经"),
    ("医学咨询", "服用方法-治疗痤疮",
     ["痤疮", "痘痘", "治疗痘痘"],
     "痤疮吃法"),
    ("医学咨询", "服用方法-用药禁忌",
     ["禁忌", "不能吃", "哺乳期", "怀孕", "备孕", "禁忌症","经期吃","抽烟","吸烟","连续不间断",
     "一次吃几片","每天吃几次","一次几片","使用方法",
     "按时吃","一次一粒","吃一颗","看您的用量","那样吃","吃21天停药7天为一个周期"],
     "咨询用药禁忌"),
    ("医学咨询", "服用方法-相互作用",
     ["相互作用", "同时吃", "其他药物", "一起吃","一块吃","可以同服","同服","影响药效"],
     "咨询药物相互作用"),
    ("医学咨询", "服用方法-其他",
     ["巧克力囊肿","紧急"],
     "咨询服用方法"),
    ("医学咨询", "不良反应-不良反应处理",
     ["副作用", "出血", "点滴出血", "褐色", "恶心", "呕吐", "头晕", "胸胀","棕色","流血","喉咙","吐痰","肚子疼",
      "不良反应", "长痘", "月经减少", "出现不良反应","及时就医","检查","连着吃"],
     "咨询副作用处理"),
    ("医学咨询", "不良反应-其他",
     ["豆腐渣分泌物", "异常出血", "其他不适"],
     "咨询其他不良反应"),
    ("购买咨询", "发货时效-咨询发货/到货",
     ["发货", "到货", "物流", "快递", "配送", "什么时候到", "到货时间","送到","还没到",
      "发货时间", "多久发货", "物流信息","加急","送达","没有货","没货"],
     "咨询发货到货时间"),
    ("购买咨询", "发货时效-其他",
     ["运输问题", "配送异常", "催促发货", "国际物流","尽快帮我送"],
     "咨询物流其他问题"),
    ("购买咨询", "价格问题-咨询优惠",
     ["优惠", "折扣", "券", "活动", "满减", "优惠券", "9折", "百亿补贴","九折劵","退差价","有卷吗",
      "省钱卡", "优惠不够", "用不了券", "赠品", "奖品", "抽奖", "e卡","怎么样买","怎样买",
      "差价", "补差价", "如果有差价", "补不补差价","全款"],
     "咨询优惠活动"),
    ("购买咨询", "价格问题-申请价保",
     ["价保", "保价", "补差价", "降价退差价"],
     "申请价保"),
    ("购买咨询", "价格问题-价格抱怨",
     ["贵", "涨价", "嫌贵", "价格高", "价格不满", "觉得贵", "变贵了",
      "之前更便宜", "价格抱怨","一个平台一个价","买的时候","涨这么多","涨太多"],
     "价格不满抱怨"),
    ("购买咨询", "价格问题-其他",
     ["价格对比", "为什么这么贵", "定价问题","价格不同","划算","结算","168一盒","原价"],
     "其他价格问题"),
    ("购买咨询", "其他",
     ["下单", "采购", "预售", "无货", "付款", "定金", "尾款", "隐私发货","地址",
      "无法下单", "下单问题", "怎么购买", "药品卡", "21片和63片区别"],
     "咨询下单购买问题"),
    ("产品质量咨询", "真伪问题-追溯码相关",
     ["追溯码", "扫码", "扫不上", "追溯码查验", "码扫不出来", "防伪码"],
     "追溯码咨询"),
    ("产品质量咨询", "真伪问题-其他",
     ["正品", "真假", "进口", "国产版", "土耳其版", "德国", "包装不同","假货","假一赔十",
      "线上线下包装不同", "是否正品", "是不是正品","真伪","假的","保真","真药","假药"],
     "正品真伪咨询"),
    ("产品质量咨询", "疑似售后-退货退款",
     ["退货", "退款", "退定金", "退货退款", "退款不同意", "拒收", "拒签","退？","可以退","暂时不需要",
      "退货地址", "要退", "退款了还发货","退了","暂不支持退换货","退一下"],
     "退货退款问题"),
    ("产品质量咨询", "疑似售后-发票相关",
     ["发票", "开票", "电子发票", "开不了发票", "修改发票"],
     "发票问题"),
    ("产品质量咨询", "疑似售后-包装外观",
     ["破损", "包装问题", "盒子坏了", "外观破损", "日期不好", "效期", "有效期",
      "生产日期", "保质期","新日期"],
     "包装外观/效期问题"),
    ("医学咨询", "基本信息-产品咨询",
     ["功效", "效果", "作用", "适应症", "成分", "避孕成功率", "医院买的不一样","做好措施",
      "从哪里开始吃", "怎么开始", "第一粒", "第一片", "哪一粒开始", "哪一片开始","服药","标签","不是减肥药","有啥区别","授权","阿里巴巴大药房",
      "几天可以同房", "第几天可以不做措施", "几天有效", "第八天有效", "连续服用7天","处方","粉色药丸","粉色药","降压药","规格",
      "第几天起效", "什么时候可以不戴套", "什么时候有效", "避孕效果","吃多久","都治","几盒","现货","有3板","有三板","都是一样的",
      "第一次吃", "月经第五天开始吃", "第五天开始吃", "服用时间","原研药","管用","链接","处方药","提供用药人的性别、年龄及当前症状",
      "粉色少了一粒", "白色药片", "吃白色期间", "少吃一片", "药片颜色","多少片","自营","长期吃","月的量","吃法是一样",
      "粉色药片", "白色的", "剩下白色", "吃完白色", "少吃一粒","吃这个","服用","餐后","睡前","睡觉前",
      "一样的吗", "药片是一样的", "包装规格不一样", "规格怎么不一样", "84片","这款药卖","使用说明书","降压药",
      "28片", "三板", "包装", "大包装", "药片一样", "规格不一样", "包装不一样","早上吃","晚上吃","中午吃",
      "长期吃", "开始吃", "吃完最后一片白片的二天","饭前","饭后","喝药","可以吃","吃可以","药片","使用吃",
      "第二盒什么时候吃", "第二盒怎么吃", "吃第二盒的时间","时间吃","吃一个","例假","吃完","药管事"],
     "咨询产品功效"),
    ("医学咨询", "服用方法-避孕",
     ["避孕", "优思明吃法", "优思悦吃法", "怎么吃", "服用方法", "忌口"],
     "咨询避孕服用方法"),
    ("其他", "其他类型",
     ["其他店", "别的平台", "对比其他品牌", "香氛", "妈富隆"],
     "非本店产品咨询"),
    ("其他", "咨询入会",
     ["入会", "会员", "注册会员", "怎么入会"],
     "会员入会咨询"),
    ("其他", "进线未咨询",
     ["未说话", "不说话", "没咨询", "离线", "弹链接", "未读", "进线没说话"],
     "进线未说话"),
]


def split_messages(dialog):
    """分离用户消息和客服消息，返回 (user_texts, agent_texts)"""
    user_texts = []
    agent_texts = []
    for line in dialog.split('\n'):
        line = line.strip()
        if not line:
            continue
        if line.startswith('[') or line.startswith('<'):
            continue
        if ': ' not in line:
            continue
        content = line.split(': ', 1)[1]
        if line.startswith('拜耳-') or line.startswith('jimi_') or line.startswith('sqjk'):
            agent_texts.append(content)
        else:
            user_texts.append(content)
    return user_texts, agent_texts


def match_text(text, full_user_text, user_msg_count):
    """对单段文本做关键词匹配，返回 (咨询类别, 问题分类) 或 None"""
    text_lower = text.lower()

    # 最高优先级：只有一句用户消息且包含"人工"
    if user_msg_count == 1 and "人工" in full_user_text.lower():
        return ("产品质量咨询", "疑似售后-其他")

    # 药品细分（用全部用户消息判断）
    full_lower = full_user_text.lower()
    if "优思明" in full_lower:
        if any(k in full_lower for k in ["推迟", "推迟月经"]):
            return ("医学咨询", "服用方法-推迟月经")
        elif any(k in full_lower for k in ["规律", "调经", "调理月经", "调节激素"]):
            return ("医学咨询", "服用方法-规律经期")
        else:
            return ("医学咨询", "服用方法-避孕")
    if "优思悦" in full_lower:
        if any(k in full_lower for k in ["推迟", "推迟月经"]):
            return ("医学咨询", "服用方法-推迟月经")
        elif any(k in full_lower for k in ["规律", "调经", "调理月经"]):
            return ("医学咨询", "服用方法-规律经期")
        elif any(k in full_lower for k in ["痤疮", "痘痘"]):
            return ("医学咨询", "服用方法-治疗痤疮")
        else:
            return ("医学咨询", "服用方法-避孕")
    if "唯散宁" in full_lower:
        return ("医学咨询", "服用方法-其他")

    for consult_cat, question_cat, keywords, detail in KEYWORD_RULES:
        if any(keyword in text_lower for keyword in keywords):
            return (consult_cat, question_cat)

    return None


def match_first(msg_list, full_user_text, user_msg_count):
    """对消息列表逐一匹配，返回第一个命中的结果"""
    for content in msg_list:
        if not content.strip():
            continue
        hit = match_text(content, full_user_text, user_msg_count)
        if hit:
            return hit
    return None


def analyze_multi_topic(dialog):
    """匹配策略：用户消息优先 → 客服消息兜底 → 全对话最终兜底，返回单话题"""
    user_texts, agent_texts = split_messages(dialog)
    if not user_texts and not agent_texts:
        return [("其他", "进线未咨询")]

    full_user_text = ' '.join(user_texts)
    user_msg_count = len(user_texts)

    # 第一轮：逐个匹配用户消息，取第一个命中
    result = match_first(user_texts, full_user_text, user_msg_count)
    if result:
        return [result]

    # 第二轮：用户消息没命中 → 匹配客服消息
    result = match_first(agent_texts, full_user_text, user_msg_count)
    if result:
        return [result]

    # 第三轮：还没命中 → 全对话兜底
    dialog_lower = dialog.lower()
    for consult_cat, question_cat, keywords, detail in KEYWORD_RULES:
        if any(keyword in dialog_lower for keyword in keywords):
            return [(consult_cat, question_cat)]
    return [("其他", "其他类型")]


def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        from copy import copy as _copy
        target_cell.font = _copy(source_cell.font)
        target_cell.border = _copy(source_cell.border)
        target_cell.fill = _copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = _copy(source_cell.protection)
        target_cell.alignment = _copy(source_cell.alignment)


def add_analysis_to_sheet(ws):
    """在结束时间右侧添加 咨询类别、问题分类 两列，支持多话题输出多行"""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {cell: idx for idx, cell in enumerate(header_row)}

    if '对话内容' not in headers or '结束时间' not in headers:
        return None
    if '咨询类别' in headers or '问题分类' in headers:
        return None

    end_time_col = headers['结束时间']
    dialog_col = headers['对话内容']

    # 在结束时间右侧插入2列
    insert_at = end_time_col + 2
    ws.insert_cols(insert_at, 2)
    dialog_col += 2

    # 表头
    ref_header = ws.cell(row=1, column=end_time_col + 1)
    for col_offset, title in enumerate(['咨询类别', '问题分类']):
        new_cell = ws.cell(row=1, column=insert_at + col_offset, value=title)
        copy_cell_style(ref_header, new_cell)

    total = 0
    for row in ws.iter_rows(min_row=2):
        if all(cell.value is None for cell in row):
            continue
        dialog = str(row[dialog_col].value) if row[dialog_col].value else ""
        topics = analyze_multi_topic(dialog) if dialog.strip() else [("其他", "进线未咨询")]
        ref_cell = row[end_time_col]
        cat, subcat = topics[0]
        cat_cell = ws.cell(row=row[0].row, column=insert_at, value=cat)
        sub_cell = ws.cell(row=row[0].row, column=insert_at + 1, value=subcat)
        copy_cell_style(ref_cell, cat_cell)
        copy_cell_style(ref_cell, sub_cell)
        total += 1

    return total


def main():
    print("=" * 60)
    print("医药电商售后数据分析器 (方案C: 用户优先)")
    print("=" * 60)

    INPUT_FILE = input("请输入Excel文件路径: ").strip().strip('"')

    if not os.path.exists(INPUT_FILE):
        print(f"错误: 文件不存在 {INPUT_FILE}")
        return

    input_dir = os.path.dirname(INPUT_FILE)
    OUTPUT_FILE = os.path.join(input_dir, "对话原因分析_AI.xlsx")

    temp_file = OUTPUT_FILE.replace('.xlsx', '_temp.xlsx')
    if os.path.exists(temp_file):
        try:
            os.remove(temp_file)
        except:
            pass

    wb = openpyxl.load_workbook(INPUT_FILE)

    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        result = add_analysis_to_sheet(ws)
        if result is None:
            print(f"  {sheet_name}: 跳过")
        else:
            total += result
            print(f"  {sheet_name}: {result} 行已分析")

    wb.save(temp_file)
    wb.close()

    try:
        os.replace(temp_file, OUTPUT_FILE)
    except:
        if os.path.exists(OUTPUT_FILE):
            os.remove(OUTPUT_FILE)
        os.rename(temp_file, OUTPUT_FILE)

    print(f"\n分析完成，共处理: {total} 行")
    print(f"文件已保存到: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
